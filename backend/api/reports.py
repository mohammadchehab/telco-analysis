from fastapi import APIRouter, Depends, HTTPException, Request
from sqlalchemy.orm import Session
from typing import List, Dict, Any, Optional
import json
from datetime import datetime
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from core.database import get_db
from services.capability_service import CapabilityService
from models.models import VendorScore
from schemas.schemas import (
    APIResponse, ReportRequest, RadarChartData, VendorComparisonData, 
    ScoreDistributionData, ChartData
)
from models.models import Capability
import logging
from models.models import Attribute, Domain
from io import BytesIO
import base64

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/reports", tags=["reports"])

@router.post("/generate", response_model=APIResponse)
async def generate_report(request: ReportRequest, db: Session = Depends(get_db)):
    """Generate comprehensive report for a capability"""
    try:
        capability = CapabilityService.get_capability(db, request.capability_id)
        if not capability:
            return APIResponse(success=False, error="Capability not found")
        
        # Generate different types of data based on report type
        if request.report_type == "vendor_comparison":
            data = CapabilityService.generate_vendor_comparison_data(db, request.capability_id)
        elif request.report_type == "radar_chart":
            data = CapabilityService.generate_radar_chart_data(db, request.capability_id)
        elif request.report_type == "score_distribution":
            data = CapabilityService.generate_score_distribution_data(db, request.capability_id)
        elif request.report_type == "comprehensive":
            # Generate all types of data for comprehensive report
            radar_data = CapabilityService.generate_radar_chart_data(db, request.capability_id)
            vendor_data = CapabilityService.generate_vendor_comparison_data(db, request.capability_id)
            distribution_data = CapabilityService.generate_score_distribution_data(db, request.capability_id)
            vendor_scores = CapabilityService.get_vendor_scores(db, capability.name)
            
            data = {
                "capability_name": capability.name,
                "generated_at": datetime.now().isoformat(),
                "radar_chart": radar_data.model_dump(),
                "vendor_comparison": vendor_data.model_dump(),
                "score_distribution": distribution_data.model_dump(),
                "vendor_scores": [score.dict() for score in vendor_scores]
            }
        else:
            return APIResponse(success=False, error="Invalid report type")
        
        # Format the response based on requested format
        if request.format == "json":
            return APIResponse(
                success=True,
                data=data.model_dump() if hasattr(data, 'model_dump') else data
            )
        elif request.format == "excel":
            excel_data = generate_excel_report(data, capability.name, request.report_type)
            return APIResponse(
                success=True,
                data={"excel_data": excel_data, "filename": f"{capability.name}_{request.report_type}_report.xlsx"}
            )
        elif request.format == "pdf":
            pdf_data = generate_pdf_report(data, capability.name, request.report_type)
            return APIResponse(
                success=True,
                data={"pdf_data": pdf_data, "filename": f"{capability.name}_{request.report_type}_report.pdf"}
            )
        else:
            return APIResponse(success=False, error="Invalid format")
            
    except Exception as e:
        return APIResponse(success=False, error=str(e))

@router.get("/{capability_id}/radar-chart", response_model=APIResponse)
async def get_radar_chart_data(capability_id: int, db: Session = Depends(get_db)):
    """Get radar chart data for a capability"""
    try:
        data = CapabilityService.generate_radar_chart_data(db, capability_id)
        return APIResponse(success=True, data=data.model_dump())
    except Exception as e:
        return APIResponse(success=False, error=str(e))

@router.get("/{capability_id}/vendor-comparison", response_model=APIResponse)
async def get_vendor_comparison_data(capability_id: int, db: Session = Depends(get_db)):
    """Get vendor comparison data for a capability"""
    try:
        data = CapabilityService.generate_vendor_comparison_data(db, capability_id)
        return APIResponse(success=True, data=data.model_dump())
    except Exception as e:
        return APIResponse(success=False, error=str(e))

@router.get("/{capability_id}/score-distribution", response_model=APIResponse)
async def get_score_distribution_data(capability_id: int, db: Session = Depends(get_db)):
    """Get score distribution data for a capability"""
    try:
        data = CapabilityService.generate_score_distribution_data(db, capability_id)
        return APIResponse(success=True, data=data.model_dump())
    except Exception as e:
        return APIResponse(success=False, error=str(e))

@router.get("/{capability_id}/comprehensive", response_model=APIResponse)
async def get_comprehensive_report(capability_id: int, db: Session = Depends(get_db)):
    """Get comprehensive report data for a capability"""
    try:
        capability = CapabilityService.get_capability(db, capability_id)
        if not capability:
            return APIResponse(success=False, error="Capability not found")
        
        # Generate all types of data
        radar_data = CapabilityService.generate_radar_chart_data(db, capability_id)
        vendor_data = CapabilityService.generate_vendor_comparison_data(db, capability_id)
        distribution_data = CapabilityService.generate_score_distribution_data(db, capability_id)
        vendor_scores = CapabilityService.get_vendor_scores(db, capability.name)
        
        comprehensive_data = {
            "capability_name": capability.name,
            "generated_at": datetime.now().isoformat(),
            "radar_chart": radar_data.model_dump(),
            "vendor_comparison": vendor_data.model_dump(),
            "score_distribution": distribution_data.model_dump(),
            "vendor_scores": [score.dict() for score in vendor_scores]
        }
        
        return APIResponse(success=True, data=comprehensive_data)
    except Exception as e:
        return APIResponse(success=False, error=str(e))

@router.get("/{capability_id}/summary", response_model=APIResponse)
async def get_capability_summary(capability_id: int, db: Session = Depends(get_db)):
    """Get summary statistics for a capability"""
    try:
        capability = CapabilityService.get_capability(db, capability_id)
        if not capability:
            return APIResponse(success=False, error="Capability not found")
        
        # Get vendor scores directly from database for summary calculations
        vendor_scores = db.query(VendorScore).filter(VendorScore.capability_id == capability_id).all()
        
        # Calculate summary statistics
        vendors = ["comarch", "servicenow", "salesforce"]
        summary_stats = {}
        
        for vendor in vendors:
            vendor_scores_list = [score for score in vendor_scores if score.vendor == vendor]
            if vendor_scores_list:
                avg_score = sum(score.score_numeric for score in vendor_scores_list) / len(vendor_scores_list)
                max_score = max(score.score_numeric for score in vendor_scores_list)
                min_score = min(score.score_numeric for score in vendor_scores_list)
                total_attributes = len(vendor_scores_list)
                
                summary_stats[vendor] = {
                    "average_score": round(avg_score, 2),
                    "max_score": max_score,
                    "min_score": min_score,
                    "total_attributes": total_attributes,
                    "score_range": f"{min_score}-{max_score}"
                }
        
        summary_data = {
            "capability_name": capability.name,
            "capability_status": capability.status,
            "total_attributes": len(vendor_scores) // len(vendors) if vendors else 0,
            "vendors_analyzed": len(vendors),
            "last_updated": datetime.now().isoformat(),
            "vendor_summaries": summary_stats
        }
        
        return APIResponse(success=True, data=summary_data)
    except Exception as e:
        return APIResponse(success=False, error=str(e))

@router.get("/{capability_id}/export/{format}", response_model=APIResponse)
async def export_capability_report(
    capability_id: int, 
    format: str, 
    report_type: str = "comprehensive",
    db: Session = Depends(get_db)
):
    """Export capability report in specified format"""
    try:
        capability = CapabilityService.get_capability(db, capability_id)
        if not capability:
            return APIResponse(success=False, error="Capability not found")
        
        # Generate report data
        if report_type == "comprehensive":
            radar_data = CapabilityService.generate_radar_chart_data(db, capability_id)
            vendor_data = CapabilityService.generate_vendor_comparison_data(db, capability_id)
            distribution_data = CapabilityService.generate_score_distribution_data(db, capability_id)
            vendor_scores = CapabilityService.get_vendor_scores(db, capability.name)
            
            data = {
                "capability_name": capability.name,
                "generated_at": datetime.now().isoformat(),
                "radar_chart": radar_data.model_dump(),
                "vendor_comparison": vendor_data.model_dump(),
                "score_distribution": distribution_data.model_dump(),
                "vendor_scores": [score.model_dump() for score in vendor_scores]
            }
        else:
            # Handle other report types
            if report_type == "vendor_comparison":
                data = CapabilityService.generate_vendor_comparison_data(db, capability_id)
            elif report_type == "radar_chart":
                data = CapabilityService.generate_radar_chart_data(db, capability_id)
            elif report_type == "score_distribution":
                data = CapabilityService.generate_score_distribution_data(db, capability_id)
            else:
                return APIResponse(success=False, error="Invalid report type")
        
        # Generate export file
        if format.lower() == "excel":
            export_data = generate_excel_report(data, capability.name, report_type)
            filename = f"{capability.name}_{report_type}_report.xlsx"
        elif format.lower() == "pdf":
            export_data = generate_pdf_report(data, capability.name, report_type)
            filename = f"{capability.name}_{report_type}_report.pdf"
        else:
            return APIResponse(success=False, error="Invalid export format")
        
        return APIResponse(
            success=True,
            data={
                "export_data": export_data,
                "filename": filename,
                "format": format,
                "report_type": report_type
            }
        )
        
    except Exception as e:
        return APIResponse(success=False, error=str(e))

@router.get("/{capability_id}/vendor-analysis", response_model=APIResponse)
async def get_vendor_analysis_data(
    capability_id: int, 
    vendors: str = "comarch,servicenow,salesforce",
    db: Session = Depends(get_db)
):
    """Get detailed vendor analysis data for comparison"""
    try:
        capability = CapabilityService.get_capability(db, capability_id)
        if not capability:
            return APIResponse(success=False, error="Capability not found")
        
        # Parse vendors from query parameter
        vendor_list = [v.strip() for v in vendors.split(",")]
        
        # Get detailed vendor analysis data
        analysis_data = CapabilityService.generate_vendor_analysis_data(db, capability_id, vendor_list)
        return APIResponse(success=True, data=analysis_data)
    except Exception as e:
        return APIResponse(success=False, error=str(e))

@router.get("/{capability_id}/vendor-analysis/export", response_model=APIResponse)
async def export_vendor_analysis(
    capability_id: int,
    vendors: str = "comarch,servicenow,salesforce",
    format: str = "excel",
    db: Session = Depends(get_db)
):
    """Export vendor analysis to Excel with detailed format"""
    try:
        capability = CapabilityService.get_capability(db, capability_id)
        if not capability:
            return APIResponse(success=False, error="Capability not found")
        
        # Parse vendors from query parameter
        vendor_list = [v.strip() for v in vendors.split(",")]
        
        # Get detailed vendor analysis data
        analysis_data = CapabilityService.generate_vendor_analysis_data(db, capability_id, vendor_list)
        
        if format == "excel":
            excel_data = generate_vendor_analysis_excel(analysis_data, capability.name, vendor_list)
            return APIResponse(
                success=True,
                data={
                    "excel_data": excel_data, 
                    "filename": f"{capability.name}_vendor_analysis.xlsx"
                }
            )
        else:
            return APIResponse(success=False, error="Only Excel format supported for vendor analysis")
            
    except Exception as e:
        return APIResponse(success=False, error=str(e))

@router.get("/vendor-analysis/export-all", response_model=APIResponse)
async def export_all_vendor_analysis(
    vendors: str = "comarch,servicenow,salesforce",
    format: str = "excel",
    db: Session = Depends(get_db)
):
    """Export vendor analysis data for all capabilities"""
    try:
        # Parse vendors from query parameter
        vendor_list = [v.strip() for v in vendors.split(",")]
        
        # Get all completed capabilities
        capabilities = db.query(Capability).filter(Capability.status == "completed").all()
        
        if not capabilities:
            return APIResponse(success=False, error="No completed capabilities found")
        
        # Generate Excel data for all capabilities
        excel_data = generate_all_capabilities_excel(capabilities, vendor_list, db)
        
        filename = f"vendor_analysis_all_capabilities_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return APIResponse(
            success=True,
            data={
                "excel_data": excel_data,
                "filename": filename
            }
        )
    except Exception as e:
        logger.error(f"Error exporting all vendor analysis: {str(e)}")
        return APIResponse(success=False, error=f"Failed to export vendor analysis: {str(e)}")

@router.get("/{capability_id}/filtered-reports", response_model=APIResponse)
async def get_filtered_reports(
    capability_id: int,
    domains: str = "",
    vendors: str = "comarch,servicenow,salesforce",
    attributes: str = "",
    db: Session = Depends(get_db)
):
    """Get filtered reports data for a capability"""
    try:
        capability = CapabilityService.get_capability(db, capability_id)
        if not capability:
            return APIResponse(success=False, error="Capability not found")
        
        # Only proceed if capability is completed
        if capability.status != "completed":
            return APIResponse(success=False, error="Capability research is not completed")
        
        # Parse filter parameters
        domain_list = [d.strip() for d in domains.split(",") if d.strip()] if domains else []
        vendor_list = [v.strip() for v in vendors.split(",") if v.strip()] if vendors else []
        attribute_list = [a.strip() for a in attributes.split(",") if a.strip()] if attributes else []
        
        # Get filtered data
        filtered_data = CapabilityService.generate_filtered_reports_data(
            db, capability_id, domain_list, vendor_list, attribute_list
        )
        
        return APIResponse(success=True, data=filtered_data)
    except Exception as e:
        return APIResponse(success=False, error=str(e))

@router.get("/{capability_id}/available-filters", response_model=APIResponse)
async def get_available_filters(capability_id: int, db: Session = Depends(get_db)):
    """Get available filter options for a capability"""
    try:
        capability = CapabilityService.get_capability(db, capability_id)
        if not capability:
            return APIResponse(success=False, error="Capability not found")
        
        # Only proceed if capability is completed
        if capability.status != "completed":
            return APIResponse(success=False, error="Capability research is not completed")
        
        # Get available filter options
        filter_options = CapabilityService.get_available_filter_options(db, capability_id)
        
        return APIResponse(success=True, data=filter_options)
    except Exception as e:
        return APIResponse(success=False, error=str(e))

def generate_excel_report(data: Any, capability_name: str, report_type: str) -> str:
    """Generate Excel report from data"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = f"{capability_name} Report"
        
        # Header
        ws['A1'] = f"{capability_name} - {report_type.replace('_', ' ').title()} Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Add data based on report type
        if report_type == "vendor_comparison":
            add_vendor_comparison_to_excel(ws, data)
        elif report_type == "radar_chart":
            add_radar_chart_to_excel(ws, data)
        elif report_type == "score_distribution":
            add_score_distribution_to_excel(ws, data)
        elif report_type == "comprehensive":
            add_comprehensive_to_excel(ws, data)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Convert to base64 for API response
        import base64
        return base64.b64encode(output.getvalue()).decode('utf-8')
        
    except Exception as e:
        raise Exception(f"Error generating Excel report: {str(e)}")

def generate_pdf_report(data: Any, capability_name: str, report_type: str) -> str:
    """Generate PDF report from data"""
    try:
        # Create PDF document
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        story = []
        
        # Add title
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30
        )
        story.append(Paragraph(f"{capability_name} - {report_type.replace('_', ' ').title()} Report", title_style))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Add data based on report type
        if report_type == "vendor_comparison":
            add_vendor_comparison_to_pdf(story, data)
        elif report_type == "radar_chart":
            add_radar_chart_to_pdf(story, data)
        elif report_type == "score_distribution":
            add_score_distribution_to_pdf(story, data)
        elif report_type == "comprehensive":
            add_comprehensive_to_pdf(story, data)
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        
        # Convert to base64 for API response
        import base64
        return base64.b64encode(output.getvalue()).decode('utf-8')
        
    except Exception as e:
        raise Exception(f"Error generating PDF report: {str(e)}")

def add_vendor_comparison_to_excel(ws, data):
    """Add vendor comparison data to Excel worksheet"""
    # Add vendor comparison table
    ws['A4'] = "Vendor Comparison"
    ws['A4'].font = Font(size=14, bold=True)
    
    # Headers
    headers = ["Attribute"] + data.vendors
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data
    for row, attribute in enumerate(data.attributes, 6):
        ws.cell(row=row, column=1, value=attribute)
        for col, vendor in enumerate(data.vendors, 2):
            score = data.scores[vendor][row-6]
            ws.cell(row=row, column=col, value=score)

def add_radar_chart_to_excel(ws, data):
    """Add radar chart data to Excel worksheet"""
    # Add radar chart data table
    ws['A4'] = "Radar Chart Data"
    ws['A4'].font = Font(size=14, bold=True)
    
    # Headers
    headers = ["Attribute"] + data.vendors
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data
    for row, attribute in enumerate(data.attributes, 6):
        ws.cell(row=row, column=1, value=attribute)
        for col, vendor_scores in enumerate(data.scores, 2):
            score = vendor_scores[row-6]
            ws.cell(row=row, column=col, value=score)

def add_score_distribution_to_excel(ws, data):
    """Add score distribution data to Excel worksheet"""
    # Add score distribution table
    ws['A4'] = "Score Distribution"
    ws['A4'].font = Font(size=14, bold=True)
    
    # Headers
    headers = ["Score Range"] + data.vendors
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data
    for row, score_range in enumerate(data.score_ranges, 6):
        ws.cell(row=row, column=1, value=score_range)
        for col, vendor in enumerate(data.vendors, 2):
            count = data.vendor_counts[vendor][row-6]
            ws.cell(row=row, column=col, value=count)

def add_comprehensive_to_excel(ws, data):
    """Add comprehensive report data to Excel worksheet"""
    # Add multiple sheets for comprehensive report
    # This is a simplified version - you can expand this based on your needs
    
    # Summary sheet
    ws['A4'] = "Capability Summary"
    ws['A4'].font = Font(size=14, bold=True)
    ws['A5'] = f"Capability: {data['capability_name']}"
    ws['A6'] = f"Generated: {data['generated_at']}"
    
    # Add vendor scores summary
    if 'vendor_scores' in data:
        ws['A8'] = "Vendor Scores Summary"
        ws['A8'].font = Font(size=12, bold=True)
        
        headers = ["Vendor", "Attribute", "Score", "Weight", "Decision"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col, value=header)
            cell.font = Font(bold=True)
        
        row = 10
        for score in data['vendor_scores']:
            ws.cell(row=row, column=1, value=score['vendor'])
            ws.cell(row=row, column=2, value=score['attribute_name'])
            ws.cell(row=row, column=3, value=score['score'])
            ws.cell(row=row, column=4, value=score['weight'])
            ws.cell(row=row, column=5, value=score['score_decision'])
            row += 1

def add_vendor_comparison_to_pdf(story, data):
    """Add vendor comparison data to PDF story"""
    story.append(Paragraph("Vendor Comparison", getSampleStyleSheet()['Heading2']))
    
    # Create table data
    table_data = [["Attribute"] + data.vendors]
    for i, attribute in enumerate(data.attributes):
        row = [attribute]
        for vendor in data.vendors:
            score = data.scores[vendor][i]
            row.append(str(score))
        table_data.append(row)
    
    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    story.append(Spacer(1, 20))

def add_radar_chart_to_pdf(story, data):
    """Add radar chart data to PDF story"""
    story.append(Paragraph("Radar Chart Data", getSampleStyleSheet()['Heading2']))
    
    # Create table data
    table_data = [["Attribute"] + data.vendors]
    for i, attribute in enumerate(data.attributes):
        row = [attribute]
        for vendor_scores in data.scores:
            score = vendor_scores[i]
            row.append(str(score))
        table_data.append(row)
    
    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    story.append(Spacer(1, 20))

def add_score_distribution_to_pdf(story, data):
    """Add score distribution data to PDF story"""
    story.append(Paragraph("Score Distribution", getSampleStyleSheet()['Heading2']))
    
    # Create table data
    table_data = [["Score Range"] + data.vendors]
    for i, score_range in enumerate(data.score_ranges):
        row = [score_range]
        for vendor in data.vendors:
            count = data.vendor_counts[vendor][i]
            row.append(str(count))
        table_data.append(row)
    
    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    story.append(Spacer(1, 20))

def add_comprehensive_to_pdf(story, data):
    """Add comprehensive report data to PDF story"""
    story.append(Paragraph("Comprehensive Report", getSampleStyleSheet()['Heading2']))
    
    # Add vendor scores summary
    if 'vendor_scores' in data:
        story.append(Paragraph("Vendor Scores Summary", getSampleStyleSheet()['Heading3']))
        
        # Create table data
        table_data = [["Vendor", "Attribute", "Score", "Weight", "Decision"]]
        for score in data['vendor_scores']:
            table_data.append([
                score['vendor'],
                score['attribute_name'],
                score['score'],
                str(score['weight']),
                score['score_decision']
            ])
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        story.append(Spacer(1, 20)) 

def generate_vendor_analysis_excel(analysis_data: dict, capability_name: str, vendors: list) -> str:
    """Generate Excel report for vendor analysis with detailed format"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = f"{capability_name} Vendor Analysis"
        
        # Header
        ws['A1'] = f"{capability_name} - Vendor Analysis Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = f"Vendors: {', '.join(vendors)}"
        
        # Define headers based on the requested format
        headers = [
            "Capability", "Domain", "Attribute", 
            "Observations (Vendor 1)", "Observations (Vendor 2)", "Observations (Vendor 3)",
            "Score Vendor 1", "Score Vendor 2", "Score Vendor 3",
            "Justification Score Vendor 1", "Justification Score Vendor 2", "Justification Score Vendor 3",
            "Evidence Vendor 1", "Evidence Vendor 2", "Evidence Vendor 3"
        ]
        
        # Adjust headers based on number of vendors
        if len(vendors) == 1:
            headers = ["Capability", "Domain", "Attribute", 
                      f"Observations ({vendors[0].title()})", 
                      f"Score {vendors[0].title()}", 
                      f"Justification Score {vendors[0].title()}", 
                      f"Evidence {vendors[0].title()}"]
        elif len(vendors) == 2:
            headers = ["Capability", "Domain", "Attribute", 
                      f"Observations ({vendors[0].title()})", f"Observations ({vendors[1].title()})",
                      f"Score {vendors[0].title()}", f"Score {vendors[1].title()}",
                      f"Justification Score {vendors[0].title()}", f"Justification Score {vendors[1].title()}",
                      f"Evidence {vendors[0].title()}", f"Evidence {vendors[1].title()}"]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        row = 6
        for item in analysis_data.get('analysis_items', []):
            col = 1
            ws.cell(row=row, column=col, value=item['capability_name']); col += 1
            ws.cell(row=row, column=col, value=item['domain_name']); col += 1
            ws.cell(row=row, column=col, value=item['attribute_name']); col += 1
            
            # Add vendor data
            for vendor in vendors:
                vendor_data = item.get('vendors', {}).get(vendor, {})
                observations = vendor_data.get('observations', [])
                if observations:
                    # Format observations as a list with types
                    obs_text = []
                    for obs in observations:
                        obs_type = obs.get('type', 'note')
                        obs_text.append(f"[{obs_type.upper()}] {obs.get('observation', '')}")
                    ws.cell(row=row, column=col, value='\n'.join(obs_text))
                else:
                    ws.cell(row=row, column=col, value='No observations available')
                col += 1
            
            for vendor in vendors:
                vendor_data = item.get('vendors', {}).get(vendor, {})
                ws.cell(row=row, column=col, value=vendor_data.get('score', '')); col += 1
            
            for vendor in vendors:
                vendor_data = item.get('vendors', {}).get(vendor, {})
                ws.cell(row=row, column=col, value=vendor_data.get('score_decision', '')); col += 1
            
            for vendor in vendors:
                vendor_data = item.get('vendors', {}).get(vendor, {})
                ws.cell(row=row, column=col, value=vendor_data.get('evidence_url', '')); col += 1
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Convert to base64 for API response
        import base64
        return base64.b64encode(output.getvalue()).decode('utf-8')
        
    except Exception as e:
        raise Exception(f"Error generating vendor analysis Excel report: {str(e)}") 

def generate_all_capabilities_excel(capabilities: list, vendors: list, db: Session) -> str:
    """Generate Excel report for all capabilities with vendor analysis"""
    try:
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create summary sheet
        summary_ws = wb.create_sheet("Summary")
        
        # Summary headers
        summary_headers = [
            "Capability", "Total Attributes", "Total Domains", 
            "Average Score (Comarch)", "Average Score (ServiceNow)", "Average Score (Salesforce)",
            "Average Score (Oracle)", "Average Score (IBM)", "Average Score (Microsoft)"
        ]
        
        for col, header in enumerate(summary_headers, 1):
            summary_ws.cell(row=1, column=col, value=header)
            summary_ws.cell(row=1, column=col).font = Font(bold=True)
        
        # Add capability summary data
        for row, capability in enumerate(capabilities, 2):
            summary_ws.cell(row=row, column=1, value=capability.name)
            
            # Get capability stats
            attributes = db.query(Attribute).filter(Attribute.capability_id == capability.id).all()
            domains = db.query(Domain).filter(Domain.capability_id == capability.id).all()
            
            summary_ws.cell(row=row, column=2, value=len(attributes))
            summary_ws.cell(row=row, column=3, value=len(domains))
            
            # Calculate average scores for each vendor
            for col, vendor in enumerate(vendors, 4):
                vendor_scores = db.query(VendorScore).filter(
                    VendorScore.capability_id == capability.id,
                    VendorScore.vendor == vendor
                ).all()
                
                if vendor_scores:
                    avg_score = sum(score.score_numeric for score in vendor_scores) / len(vendor_scores)
                    summary_ws.cell(row=row, column=col, value=round(avg_score, 2))
                else:
                    summary_ws.cell(row=row, column=col, value="N/A")
        
        # Create detailed analysis sheets for each capability
        for capability in capabilities:
            try:
                # Generate detailed data for this capability
                analysis_data = CapabilityService.generate_vendor_analysis_data(db, capability.id, vendors)
                
                # Create sheet for this capability
                sheet_name = capability.name[:31] if len(capability.name) <= 31 else capability.name[:28] + "..."
                ws = wb.create_sheet(sheet_name)
                
                # Headers for detailed analysis
                headers = ["Capability", "Domain", "Attribute"]
                for i, vendor in enumerate(vendors, 1):
                    headers.extend([
                        f"Observations (Vendor {i})",
                        f"Score (Vendor {i})",
                        f"Justification (Vendor {i})",
                        f"Evidence (Vendor {i})"
                    ])
                
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                    ws.cell(row=1, column=col).font = Font(bold=True)
                
                # Add detailed data
                for row, item in enumerate(analysis_data['analysis_items'], 2):
                    ws.cell(row=row, column=1, value=item['capability_name'])
                    ws.cell(row=row, column=2, value=item['domain_name'])
                    ws.cell(row=row, column=3, value=item['attribute_name'])
                    
                    col = 4
                    for vendor in vendors:
                        vendor_data = item['vendors'].get(vendor, {})
                        observations = vendor_data.get('observations', [])
                        if observations:
                            # Format observations as a list with types
                            obs_text = []
                            for obs in observations:
                                obs_type = obs.get('type', 'NOTE')
                                obs_text.append(f"[{obs_type}] {obs.get('observation', '')}")
                            ws.cell(row=row, column=col, value='\n'.join(obs_text))
                        else:
                            ws.cell(row=row, column=col, value='No observations available')
                        ws.cell(row=row, column=col + 1, value=vendor_data.get('score', 'N/A'))
                        ws.cell(row=row, column=col + 2, value=vendor_data.get('score_decision', 'N/A'))
                        ws.cell(row=row, column=col + 3, value=vendor_data.get('evidence_url', 'N/A'))
                        col += 4
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                    
            except Exception as e:
                logger.error(f"Error generating sheet for capability {capability.name}: {str(e)}")
                continue
        
        # Auto-adjust summary sheet column widths
        for column in summary_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            summary_ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Convert to base64
        excel_bytes = output.getvalue()
        excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')
        
        return excel_base64
        
    except Exception as e:
        logger.error(f"Error generating all capabilities Excel: {str(e)}")
        raise e 