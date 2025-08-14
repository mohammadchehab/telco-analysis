#!/usr/bin/env python3
import os
import argparse
from pathlib import Path
from typing import Dict, List, Optional, Any, Set

import pandas as pd
import openpyxl
from sqlalchemy import create_engine, text, bindparam

VENDORS = ["salesforce", "servicenow", "comarch"]

def resolve_default_db_url() -> str:
	# Prefer env; else fall back to SQLite DB inside ./backend
	db_url = os.getenv("DATABASE_URL")
	if db_url:
		return db_url
	backend_db = Path.cwd() / "backend" / "telco_analysis.db"
	return f"sqlite:///{backend_db}"

def make_engine(db_url: Optional[str]):
	url = db_url or resolve_default_db_url()
	engine = create_engine(
		url,
		connect_args={"check_same_thread": False} if url.startswith("sqlite") else {}
	)
	return engine

def fetch_base_rows(engine) -> pd.DataFrame:
	# One wide row per capability/domain/attribute with vendor-specific scores, decisions, and vs IDs
	sql = """
SELECT
	c.id AS capability_id,
	c.name AS capability_name,
	d.id AS domain_id,
	d.domain_name,
	a.id AS attribute_id,
	a.attribute_name,
	a.definition,
	a.importance,

	MAX(CASE WHEN lower(vs.vendor)='salesforce' THEN vs.score_numeric END) AS sf_score,
	MAX(CASE WHEN lower(vs.vendor)='servicenow' THEN vs.score_numeric END) AS snow_score,
	MAX(CASE WHEN lower(vs.vendor)='comarch'    THEN vs.score_numeric END) AS comarch_score,

	MAX(CASE WHEN lower(vs.vendor)='salesforce' THEN vs.score_decision END) AS sf_decision,
	MAX(CASE WHEN lower(vs.vendor)='servicenow' THEN vs.score_decision END) AS snow_decision,
	MAX(CASE WHEN lower(vs.vendor)='comarch'    THEN vs.score_decision END) AS comarch_decision,

	MAX(CASE WHEN lower(vs.vendor)='salesforce' THEN vs.id END) AS sf_vs_id,
	MAX(CASE WHEN lower(vs.vendor)='servicenow' THEN vs.id END) AS snow_vs_id,
	MAX(CASE WHEN lower(vs.vendor)='comarch'    THEN vs.id END) AS comarch_vs_id

FROM capabilities AS c
LEFT JOIN domains AS d
	ON d.capability_id = c.id
JOIN attributes AS a
	ON a.capability_id = c.id
LEFT JOIN vendor_scores AS vs
	ON vs.attribute_id = a.id
GROUP BY
	c.id, c.name, d.id, d.domain_name,
	a.id, a.attribute_name, a.definition, a.importance
ORDER BY
	c.id, d.id, a.attribute_name
"""
	with engine.connect() as conn:
		df = pd.read_sql_query(sql, conn)
	return df

def fetch_observations_map(engine, vendor_score_ids: Set[int]) -> Dict[int, List[Dict[str, Any]]]:
	if not vendor_score_ids:
		return {}
	stmt = (
		text("""
			SELECT vendor_score_id, observation_type, observation
			FROM vendor_score_observations
			WHERE vendor_score_id IN :ids
		""")
		.bindparams(bindparam("ids", expanding=True))
	)
	with engine.connect() as conn:
		rows = conn.execute(stmt, {"ids": list(vendor_score_ids)}).mappings().all()
	result: Dict[int, List[Dict[str, Any]]] = {}
	for r in rows:
		vsid = r["vendor_score_id"]
		otype = str(r["observation_type"]).lower() if r["observation_type"] is not None else ""
		result.setdefault(vsid, []).append(
			{"type": otype, "text": r["observation"] or ""}
		)
	return result

def gather_text(observations: List[Dict[str, Any]], target_type: str) -> str:
	target_type = target_type.lower()
	items = [o["text"].strip() for o in observations if (o.get("type") or "").lower() == target_type and (o.get("text") or "").strip()]
	return "; ".join(items)

def make_justification(observations: List[Dict[str, Any]], score: Optional[float], decision: Optional[str]) -> str:
	lines = []
	# Add all observation types in order
	for obs_type in ["STRENGTH", "WEAKNESS", "GAP", "FEATURE", "LIMITATION", "ADVANTAGE", "DISADVANTAGE", "NOTE"]:
		text = gather_text(observations or [], obs_type.lower())
		lines.append(f"{obs_type}: {text}" if text else f"{obs_type}: -")
	
	# Add score at the end
	if score is not None and decision:
		lines.append(f"SCORE: scored {score} because {decision}")
	elif score is not None:
		lines.append(f"SCORE: scored {score}")
	elif decision:
		lines.append(f"SCORE: {decision}")
	else:
		lines.append("SCORE: -")
	
	return "\n".join(lines)

OBSERVATION_TYPES = ["STRENGTH", "WEAKNESS", "GAP", "FEATURE", "LIMITATION", "ADVANTAGE", "DISADVANTAGE", "NOTE"]
VENDORS = ["SF", "SNOW", "COMARCH"]

def build_export(df: pd.DataFrame, obs_map: Dict[int, List[Dict[str, Any]]]) -> pd.DataFrame:
	# Compute per-vendor justification
	out_rows = []
	for _, r in df.iterrows():
		sf_vs_id = r.get("sf_vs_id")
		snow_vs_id = r.get("snow_vs_id")
		comarch_vs_id = r.get("comarch_vs_id")

		sf_obs = obs_map.get(int(sf_vs_id)) if pd.notna(sf_vs_id) else []
		snow_obs = obs_map.get(int(snow_vs_id)) if pd.notna(snow_vs_id) else []
		comarch_obs = obs_map.get(int(comarch_vs_id)) if pd.notna(comarch_vs_id) else []

		# Base data
		base_cols = {
			"capability_id": r.get("capability_id"),
			"capability_name": r.get("capability_name"),
			"domain_id": r.get("domain_id"),
			"domain_name": r.get("domain_name"),
			"attribute_id": r.get("attribute_id"),
			"attribute_name": r.get("attribute_name"),
			"definition": r.get("definition"),
			"importance": r.get("importance"),
		}

		# Scores
		score_cols = {
			"SF_score": r.get("sf_score"),
			"SNOW_score": r.get("snow_score"),
			"COMARCH_score": r.get("comarch_score"),
		}

		# Full justification columns (with bold labels)
		vendor_obs = {
			"SF": sf_obs,
			"SNOW": snow_obs,
			"COMARCH": comarch_obs
		}
		
		just_cols = {
			"SF_justification": make_justification(sf_obs, r.get("sf_score"), r.get("sf_decision")),
			"SNOW_justification": make_justification(snow_obs, r.get("snow_score"), r.get("snow_decision")),
			"COMARCH_justification": make_justification(comarch_obs, r.get("comarch_score"), r.get("comarch_decision"))
		}

		# Individual observation columns
		obs_cols = {}
		for vendor in VENDORS:
			for obs_type in OBSERVATION_TYPES:
				col_name = f"{vendor}_{obs_type}"
				obs = vendor_obs[vendor]
				obs_cols[col_name] = gather_text(obs, obs_type.lower())

		# Combine all columns in desired order
		row_data = {
			**base_cols,          # Basic info (capability_id, name, etc.)
			**score_cols,         # Vendor scores
			**just_cols,          # Full justification columns right after scores
			**obs_cols            # Individual observation columns at the end
		}

		out_rows.append(row_data)
	return pd.DataFrame(out_rows)

def main():
	parser = argparse.ArgumentParser(description="Export capability/domain/attribute vendor scores with formatted Justification.")
	parser.add_argument("--db", help="SQLAlchemy DATABASE_URL; defaults to env DATABASE_URL or sqlite:///./backend/telco_analysis.db")
	parser.add_argument("--out", default="vendor_scores_export.csv", help="Output file path (.csv or .xlsx)")
	args = parser.parse_args()

	engine = make_engine(args.db)
	base_df = fetch_base_rows(engine)

	vs_id_cols = ["sf_vs_id", "snow_vs_id", "comarch_vs_id"]
	ids: Set[int] = {
		int(vsid) for vsid in pd.unique(base_df[vs_id_cols].values.ravel("K"))
		if pd.notna(vsid)
	}
	obs_map = fetch_observations_map(engine, ids)

	export_df = build_export(base_df, obs_map)

	out_path = Path(args.out)
	if out_path.suffix.lower() == ".xlsx":
		# Export to Excel with formatting
		with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
			export_df.to_excel(writer, index=False, sheet_name='Vendor Scores')
			
			# Get the workbook and worksheet
			workbook = writer.book
			worksheet = writer.sheets['Vendor Scores']
			
			# Create bold font style
			from openpyxl.styles import Font
			bold_font = Font(bold=True)
			
			from openpyxl.styles import Font, PatternFill, Alignment
			
			# Format headers
			header_font = Font(bold=True, size=11)
			for col in range(1, worksheet.max_column + 1):
				cell = worksheet.cell(row=1, column=col)
				cell.font = header_font
				
				# Color-code vendor-specific columns
				col_name = export_df.columns[col-1]
				if any(col_name.startswith(f"{v}_") for v in VENDORS):
					if col_name.startswith("SF_"):
						color = "E3F2FD"  # Light blue
					elif col_name.startswith("SNOW_"):
						color = "F3E5F5"  # Light purple
					elif col_name.startswith("COMARCH_"):
						color = "E8F5E9"  # Light green
					cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

			from openpyxl.styles import Font, PatternFill, Alignment
			from openpyxl.utils import get_column_letter
			
			# Format justification columns
			just_cols = ['SF_justification', 'SNOW_justification', 'COMARCH_justification']
			for col_name in just_cols:
				if col_name in export_df.columns:
					col_idx = export_df.columns.get_loc(col_name) + 1
					
					# Format header
					header_cell = worksheet.cell(row=1, column=col_idx)
					header_cell.font = Font(bold=True, size=11)
					if col_name.startswith("SF_"):
						header_cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
					elif col_name.startswith("SNOW_"):
						header_cell.fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
					elif col_name.startswith("COMARCH_"):
						header_cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
					
					# Process each row
					for row in range(2, len(export_df) + 2):
						cell = worksheet.cell(row=row, column=col_idx)
						
						if cell.value:
							from openpyxl.cell.text import InlineFont
							from openpyxl.cell.rich_text import TextBlock, CellRichText
							
							lines = cell.value.split('\n')
							rich_text = CellRichText()
							
							for i, line in enumerate(lines):
								if ':' in line:
									label, content = line.split(':', 1)
									# Add bold label with colon
									rich_text.append(TextBlock(InlineFont(b=True), f"{label}:"))
									# Add normal content
									rich_text.append(content)
									# Add newline if not last line
									if i < len(lines) - 1:
										rich_text.append('\n')
							
							# Set the formatted text
							cell.value = rich_text
							
							# Apply cell formatting
							cell.alignment = Alignment(wrapText=True, vertical='top')
							
							# Set width to accommodate content
							max_line_length = max(len(line) for line in lines) if lines else 0
							worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_line_length + 2, 100)
			
			# Add radar chart comparing vendor scores
			from openpyxl.chart import RadarChart, Reference
			
			# Create a new sheet for charts
			chart_sheet = workbook.create_sheet(title='Score Analysis')
			
			# Prepare data for radar chart - average scores per capability
			capability_scores = {}
			for row in range(2, len(export_df) + 2):
				cap_name = worksheet.cell(row=row, column=2).value  # capability_name column
				if cap_name not in capability_scores:
					capability_scores[cap_name] = {
						'SF': [], 'SNOW': [], 'COMARCH': []
					}
				
				sf_score = worksheet.cell(row=row, column=export_df.columns.get_loc('SF_score') + 1).value
				snow_score = worksheet.cell(row=row, column=export_df.columns.get_loc('SNOW_score') + 1).value
				comarch_score = worksheet.cell(row=row, column=export_df.columns.get_loc('COMARCH_score') + 1).value
				
				if sf_score is not None and str(sf_score).strip():
					try:
						capability_scores[cap_name]['SF'].append(float(sf_score))
					except (ValueError, TypeError):
						pass
				
				if snow_score is not None and str(snow_score).strip():
					try:
						capability_scores[cap_name]['SNOW'].append(float(snow_score))
					except (ValueError, TypeError):
						pass
				
				if comarch_score is not None and str(comarch_score).strip():
					try:
						capability_scores[cap_name]['COMARCH'].append(float(comarch_score))
					except (ValueError, TypeError):
						pass
			
			# Calculate averages and write to chart sheet
			chart_sheet.cell(row=1, column=1, value='Capability')
			chart_sheet.cell(row=1, column=2, value='Salesforce')
			chart_sheet.cell(row=1, column=3, value='ServiceNow')
			chart_sheet.cell(row=1, column=4, value='Comarch')
			
			row = 2
			for cap_name, scores in capability_scores.items():
				chart_sheet.cell(row=row, column=1, value=cap_name)
				chart_sheet.cell(row=row, column=2, value=sum(scores['SF'])/len(scores['SF']) if scores['SF'] else 0)
				chart_sheet.cell(row=row, column=3, value=sum(scores['SNOW'])/len(scores['SNOW']) if scores['SNOW'] else 0)
				chart_sheet.cell(row=row, column=4, value=sum(scores['COMARCH'])/len(scores['COMARCH']) if scores['COMARCH'] else 0)
				row += 1
			
			# Create and style the radar chart
			chart = RadarChart()
			chart.type = 'filled'  # or 'standard' for lines only
			chart.style = 26  # Choose a style number between 1-48
			chart.title = 'Vendor Scores by Capability'
			
			# Add data to chart
			labels = Reference(chart_sheet, min_col=1, min_row=2, max_row=row-1)
			data = Reference(chart_sheet, min_col=2, max_col=4, min_row=1, max_row=row-1)
			
			chart.add_data(data, titles_from_data=True)
			chart.set_categories(labels)
			
			# Make it larger
			chart.height = 15  # Default is 7.5
			chart.width = 15   # Default is 15
			
			# Add to chart sheet
			chart_sheet.add_chart(chart, "F2")
			
			# Set column widths and text wrapping
			for col in worksheet.columns:
				col_letter = col[0].column_letter
				col_name = worksheet.cell(row=1, column=col[0].column).value
				
				# Set specific widths based on column type
				if col_name in ['capability_name', 'domain_name', 'attribute_name', 'definition']:
					# These columns can be wider since they contain descriptive text
					worksheet.column_dimensions[col_letter].width = 40
				elif col_name in ['SF_justification', 'SNOW_justification', 'COMARCH_justification']:
					# Justification columns should be wide enough for formatted text
					worksheet.column_dimensions[col_letter].width = 60
				elif any(col_name.startswith(f"{v}_") for v in VENDORS):
					# Individual observation columns (STRENGTH, WEAKNESS, etc.)
					worksheet.column_dimensions[col_letter].width = 50
				elif col_name in ['SF_score', 'SNOW_score', 'COMARCH_score']:
					# Score columns can be narrow
					worksheet.column_dimensions[col_letter].width = 10
				else:
					# Other columns get auto-width with a reasonable max
					max_length = 0
					for cell in col:
						try:
							max_length = max(max_length, len(str(cell.value or '')))
						except:
							pass
					worksheet.column_dimensions[col_letter].width = min(max_length + 2, 30)
				
				# Apply text wrapping to all cells in text-heavy columns
				if col_name in ['capability_name', 'domain_name', 'attribute_name', 'definition'] or \
				   col_name in ['SF_justification', 'SNOW_justification', 'COMARCH_justification'] or \
				   any(col_name.startswith(f"{v}_") for v in VENDORS):
					for cell in col:
						if cell.value:
							cell.alignment = Alignment(wrapText=True, vertical='top')
							
							# Adjust row height based on content and column width
							if cell.value:
								# Calculate how many lines this cell will need
								import math
								col_width = worksheet.column_dimensions[col_letter].width
								text_length = len(str(cell.value))
								num_lines = math.ceil(text_length / (col_width * 1.2))  # 1.2 chars per width unit
								
								# Get current row height and update if needed
								row_height = worksheet.row_dimensions.get(cell.row, None)
								if row_height is None:
									worksheet.row_dimensions[cell.row] = openpyxl.worksheet.dimensions.RowDimension(
										worksheet, cell.row, height=15 * max(num_lines, 1)
									)
								else:
									row_height.height = max(row_height.height or 15, 15 * max(num_lines, 1))
	else:
		export_df.to_csv(out_path, index=False)
	print(f"Wrote {len(export_df)} rows to {out_path}")

if __name__ == "__main__":
	main()